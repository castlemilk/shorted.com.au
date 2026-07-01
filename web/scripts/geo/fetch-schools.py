#!/usr/bin/env python3
"""Fetch the ACARA national school-location list (all sectors) for the suburb
school-sector split, and stage it as .staging/schools-sector.json for
join-amenities.mjs.

Source: ACARA "School Location <year>" (Australian Curriculum, Assessment and
Reporting Authority) — the national list of every school with sector
(Government/Catholic/Independent), type (Primary/Secondary/Combined) and
latitude/longitude. ~11k schools, all states, all geocoded. Served from Azure
Blob (no WAF; a browser UA is enough). Attribution: "Source: ACARA".

Note: ACARA's School Location download is governed by ACARA's My School terms of
use (not CC-BY); use is per the project owner's decision. Only sector/type/
location are read here — never ICSEA or enrolment (which are in the separate
School Profile files and are excluded).
"""
import json
import os
import urllib.request
from io import BytesIO

import openpyxl

YEAR = os.environ.get("ACARA_YEAR", "2025")
URL = (
    "https://dataandreporting.blob.core.windows.net/anrdataportal/"
    f"Data-Access-Program/School%20Location%20{YEAR}.xlsx"
)
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"

STAGING = os.path.join(os.path.dirname(__file__), ".staging")
os.makedirs(STAGING, exist_ok=True)
OUT = os.path.join(STAGING, "schools-sector.json")

SECTOR = {"Government": "gov", "Catholic": "catholic", "Independent": "independent"}


def fetch_xlsx(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=90) as resp:
        return resp.read()


def main() -> None:
    print(f"fetching ACARA School Location {YEAR} …")
    raw = fetch_xlsx(URL)
    if raw[:2] != b"PK":
        raise SystemExit(f"ACARA fetch did not return an xlsx ({len(raw)} bytes)")
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "location" in s.lower()), wb.sheetnames[-1])
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(rows)]
    ci = {h.lower(): i for i, h in enumerate(hdr)}
    iSec, iType = ci["school sector"], ci["school type"]
    iLat, iLon = ci["latitude"], ci["longitude"]

    out = []
    skipped = 0
    for r in rows:
        try:
            lat, lon = float(r[iLat]), float(r[iLon])
        except (TypeError, ValueError):
            skipped += 1
            continue
        sector = SECTOR.get(str(r[iSec]).strip())
        if not sector:
            skipped += 1
            continue
        t = str(r[iType]).strip().lower()  # Primary | Secondary | Combined
        offers_primary = t in ("primary", "combined")
        offers_secondary = t in ("secondary", "combined")
        out.append({"lon": lon, "lat": lat, "sector": sector,
                    "p": 1 if offers_primary else 0, "s": 1 if offers_secondary else 0})

    with open(OUT, "w") as f:
        json.dump(out, f)
    by_sector = {}
    for s in out:
        by_sector[s["sector"]] = by_sector.get(s["sector"], 0) + 1
    print(f"wrote {OUT}: {len(out)} schools ({by_sector}); skipped {skipped}")


if __name__ == "__main__":
    main()
